import json
from multiprocessing.spawn import get_command_line
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter




if __name__=="__main__":
    #json file
    json_str=json.load(open('testrun.json'))
    #config file
    work_book=load_workbook('config.xlsx')
    table=work_book["Table Name"]
    column=work_book["Column Name"]
    table_json={}
    row_count=2
    while table['A'+str(row_count)].value!=None:
        table_json.update({table['A'+str(row_count)].value:{
            "column":{},
            "rearrange column":table['C'+str(row_count)].value,
            "font":table['D'+str(row_count)].value,
            "search bar":table['E'+str(row_count)].value,
            "column visibility":table['F'+str(row_count)].value,
            "Pagination":table['G'+str(row_count)].value,
            "Show count":table['H'+str(row_count)].value,
            "Show filter":table['I'+str(row_count)].value,
            "Ag grid":table['J'+str(row_count)].value,
            "Period filter":table['K'+str(row_count)].value
            }})
        row_count+=1
    row_count=2
    for table_name in table_json:
        row_count=2
        while column['A'+str(row_count)].value!=None:
            if column['A'+str(row_count)].value==table_name:
                table_json[table_name]['column'].update({column['B'+str(row_count)].value:{
                "column value":column['C'+str(row_count)].value,
                "hyperlinked":column['D'+str(row_count)].value,
                "sort":column['E'+str(row_count)].value,
                "Search":column['F'+str(row_count)].value,
                "ag grid":column['G'+str(row_count)].value,
                "filter":column['H'+str(row_count)].value,
                "frozen column":column['I'+str(row_count)].value,
                "validation count":column['J'+str(row_count)].value
                }})
                value_count=0
                for value_count in range(0,int(table_json[table_name]['column'][column['B'+str(row_count)].value]["validation count"])):
                    table_json[table_name]['column'][column['B'+str(row_count)].value].update({"validation"+str(value_count+1):column[chr(75+value_count)+str(row_count)].value})
                    value_count+=1   
            row_count+=1

    allow =work_book['Allow']
    config_json={"button":{},"check box":{},"icon":{},"link":{},"tab":{}}
    conf={}
    row_count=2
    while allow['Q'+str(row_count)].value!=None:
        conf.update({allow['Q'+str(row_count)].value:{"message":allow['R'+str(row_count)].value,"yes":allow['S'+str(row_count)].value,"No":allow['T'+str(row_count)].value}})
        row_count+=1
    row_count=2
    while allow['A'+str(row_count)].value!=None:
        config_json["button"].update({allow['A'+str(row_count)].value:allow['B'+str(row_count)].value})
        row_count+=1
    row_count=2
    while allow['D'+str(row_count)].value!=None:
        config_json["check box"].update({allow['D'+str(row_count)].value:{"checked":allow['E'+str(row_count)].value,"unchecked":allow['F'+str(row_count)].value}})
        row_count+=1
    row_count=2
    while allow['H'+str(row_count)].value!=None:
        config_json["icon"].update({allow['H'+str(row_count)].value:allow['I'+str(row_count)].value})
        row_count+=1
    row_count=2
    while allow['K'+str(row_count)].value!=None:
        config_json["link"].update({allow['K'+str(row_count)].value:allow['K'+str(row_count)].value})
        row_count+=1
    row_count=2
    while allow['N'+str(row_count)].value!=None:
        config_json["tab"].update({allow['N'+str(row_count)].value:allow['O'+str(row_count)].value})
        row_count+=1
    Show_filter=work_book['Show filter']
    ag_grid=work_book['Ag grid']
    general=work_book['General']
    #Test case sheet
    work_book2=load_workbook('test case.xlsx')
    output_sheet=work_book2['Sheet1']
    row_count=2
    for senario in json_str:
        output_sheet['H'+str(row_count)]=senario
        for submodule in json_str[senario]:
            output_sheet['E'+str(row_count)]=submodule
            for funct in json_str[senario][submodule]:
                output_sheet['F'+str(row_count)]=funct
                for field in json_str[senario][submodule][funct]:
                    output_sheet['G'+str(row_count)]=field
                    for value in json_str[senario][submodule][funct][field]:
                        if value == 'allow':
                            for allow_type in json_str[senario][submodule][funct][field][value]:
                                if allow_type=='module':
                                    output_sheet['O'+str(row_count)]='Check whether the application allow to click '+json_str[senario][submodule][funct][field][value][allow_type]+' module'
                                    output_sheet['Q'+str(row_count)]='Should display '+json_str[senario][submodule][funct][field][value][allow_type]+' page'
                                    row_count=row_count+1
                                elif allow_type=='checkbox':
                                    output_sheet['O'+str(row_count)]='Check whether the application allow to check '+json_str[senario][submodule][funct][field][value][allow_type]+' checkbox'
                                    output_sheet['Q'+str(row_count)]='Should display '+config_json['check box'][json_str[senario][submodule][funct][field][value][allow_type]]['checked']+' page'
                                    row_count=row_count+1
                                    output_sheet['O'+str(row_count)]='Check whether the application allow to uncheck '+json_str[senario][submodule][funct][field][value][allow_type]+' checkbox'
                                    output_sheet['Q'+str(row_count)]='Should display '+config_json['check box'][json_str[senario][submodule][funct][field][value][allow_type]]['unchecked']+' page'
                                    row_count+=1
                                elif allow_type=='icon':
                                    output_sheet['O'+str(row_count)]='Check whether the application allow to click '+json_str[senario][submodule][funct][field][value][allow_type]+' icon'
                                    output_sheet['Q'+str(row_count)]='Should '+config_json['icon'][json_str[senario][submodule][funct][field][value][allow_type]]+' while clicking'
                                    row_count+=1
                                elif allow_type=='link':
                                    output_sheet['O'+str(row_count)]='Check whether the application display a clickable "'+json_str[senario][submodule][funct][field][value][allow_type]+'" link'
                                    output_sheet['Q'+str(row_count)]='Should '+config_json['link'][json_str[senario][submodule][funct][field][value][allow_type]]+' while cliking'
                                    row_count+=1
                                elif allow_type=='button':
                                    output_sheet['O'+str(row_count)]='Check whether the application display a clickable "'+json_str[senario][submodule][funct][field][value][allow_type]+'" button'
                                    output_sheet['Q'+str(row_count)]='Should '+config_json['button'][json_str[senario][submodule][funct][field][value][allow_type]]+' while clicking'
                                    row_count+=1
                                elif allow_type=='popup page':
                                    print('popup page')
                                elif allow_type=='confirmation':
                                    output_sheet['O'+str(row_count)]='Check whether the application display "'+conf[json_str[senario][submodule][funct][field][value][allow_type]]['message']+'" confirmation msg with Yes and No button'
                                    output_sheet['Q'+str(row_count)]='Should display "'+conf[json_str[senario][submodule][funct][field][value][allow_type]]['message']+'" confirmation msg with Yes and No button'
                                    row_count+=1
                                    output_sheet['O'+str(row_count)]='Verify the application Allow to click Yes button'
                                    output_sheet['Q'+str(row_count)]='Should '+conf[json_str[senario][submodule][funct][field][value][allow_type]]['yes']
                                    row_count+=1
                                    output_sheet['O'+str(row_count)]='Verify the application Allow to click No button'
                                    output_sheet['Q'+str(row_count)]='Should '+conf[json_str[senario][submodule][funct][field][value][allow_type]]['No']
                                    row_count+=1
                                elif allow_type=='tab':
                                    output_sheet['O'+str(row_count)]='Check whether the application allow to click '+json_str[senario][submodule][funct][field][value][allow_type]+' tab'
                                    output_sheet['Q'+str(row_count)]='Should display '+config_json['tab'][json_str[senario][submodule][funct][field][value][allow_type]]
                                    row_count+=1
                        elif value=='table':
                            print('table')
                            print(json_str[senario][submodule][funct][field][value])
                            for column_name in table_json[json_str[senario][submodule][funct][field][value]]['column']:
                                print(column_name)
                            
                        elif value=='general':
                            print('general')
    work_book2.save('test case.xlsx')